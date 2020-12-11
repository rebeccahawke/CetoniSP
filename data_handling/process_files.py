"""
A collection of assorted helper functions to collate and process data files from Igor Pro.
"""
import os
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import load_workbook, Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

from data_handling.plot_data import fit_linear, fit_quadratic, fit_sinusoid


def convert_cal_rf(data):
    # copied from RFCounter class because it wasn't letting me import it ?!
    # Calibration at 20/1/2020: y = 0.001135274x - 25.139001220
    # where y = height and x is RFC value in Hz (typically 25-35 kHz)
    # this calibration was from a two-point measurement

    # Calibration from 6/11/2020 using dial gauge and comparing several runs:
    pars = [-7.67043811e-12,  5.66480566e-07, -1.27538994e-02,  8.91495880e+01]
    # quartic coeffs = [2.30559894e-15, - 2.39545023e-10, 9.29559827e-06, - 1.58540497e-01,   1.00054383e+03]
    # typical error is around 20 um

    heights = []

    for raw_val in data:
        x = float(raw_val)
        height = pars[0]*x**3 + pars[1]*x**2 + pars[2]*x + pars[3]
        heights.append(height)

    return heights


def convert_cal_lvdt(data):
    # where y = height in mm and x is LVDT value in V (typically 1 - 7 V)
    # First quadratic fit from RFCounter calibration values as at 27/10/2020: a = -0.03542046; b = 1.07974847; c = -0.39624465
    # see ...MSL Kibble Balance\_p_PressureManifoldConsiderations\Flow and Pressure control\SyringePumpTests\20201023_RFC-LVDT_IgorData for raw data
    # Second quadratic fit from quick measurement at end of day of data collection 9/11/2020, using RFC as calibration reference

    heights = []

    for raw_val in data:
        a = -0.03315761
        b = 1.15672777
        c = -1.58507681

        height = a*float(raw_val)**2 + b*float(raw_val) + c
        heights.append(height)

    return heights


def get_all_fnames(mydir, pattern, endpattern=None):
    files = []

    for file in os.listdir(mydir):
        if file.startswith(pattern):
            if endpattern is None:
                files.append(file)
            else:
                if file.endswith(endpattern):
                    files.append(file)

    return files


def get_all_xlsx_fnames(mydir):
    files = []
    igor_data = []

    for file in os.listdir(mydir):
        if file.endswith(".xlsx"):
            files.append(file)
        if "dP_volt" in file:
            igor_data.append(file)

    return files, igor_data


def get_all_LVDTxlsx_fnames(mydir):
    files = []

    for file in os.listdir(mydir):
        if file.endswith("_LVDT.xlsx"):
            files.append(file)

    return files


def collate_sp_rfc_lvdt(folder):

    spfs = sorted(get_all_fnames(folder, "Step_", endpattern=".csv"), key=lambda x: x.split("_")[3])
    # for i in range(len(spfs)):
    #     print(spfs[i])

    rfcfs = sorted(get_all_fnames(folder, "RF-data_", endpattern=".csv"))
    # for i in range(len(rfcfs)):
    #     print(rfcfs[i])

    lvdtfs = sorted(get_all_fnames(folder, "LVDT-data_", endpattern=".csv"))
    # for i in range(len(lvdtfs)):
    #     print(lvdtfs[i])

    for i in range(len(spfs)):
        print(spfs[i], lvdtfs[i], rfcfs[i])
        sp_df = pd.read_csv(os.path.join(folder, spfs[i]))
        rfc_df = pd.read_csv(os.path.join(folder, rfcfs[i]))
        lvdt_df = pd.read_csv(os.path.join(folder, lvdtfs[i]))  # updated for collection using Python not Igor Pro
        lvdt_df.rename(columns={'Timestamp': 'Time (s)'}, inplace=True)
        lvdt_ht = pd.DataFrame(convert_cal_lvdt(lvdt_df["Voltage (V)"]), columns=["LVDT (mm)"])
        h_data = pd.concat([rfc_df, lvdt_df, lvdt_ht], axis=1)

        savepath = os.path.join(folder, spfs[i].strip(".csv")+ "_all.xlsx")

        wb = Workbook()
        sheet1 = wb.active
        sheet1.title = "SP data"
        for r in dataframe_to_rows(sp_df, index=True, header=True):
            sheet1.append(r)
        sheet2 = wb.create_sheet("RFC data")
        for r in dataframe_to_rows(h_data, index=True, header=True):
            sheet2.append(r)

        wb.save(filename=savepath)

        plot_timeseries_data(sp_df, h_data, savepath=savepath.strip("xlsx")+"png")


def read_in_data(folder, fname):
    f1 = os.path.join(folder, fname)

    xls = pd.ExcelFile(f1)
    df1 = pd.read_excel(xls, 'SP data')
    df2 = pd.read_excel(xls, 'RFC data')

    return df1, df2


def plot_sp_rfc_data(SP_csv, RFC_csv, savepath=None):

    sp_data = pd.read_csv(SP_csv)
    rfc_data = pd.read_csv(RFC_csv)

    fig, (ax1, ax2) = plt.subplots(2, sharex=True)
    # fig.suptitle('Piston height over time')
    ax1.set_ylabel('Height (mm)')
    # ax1.set_title('From Capacitor')
    ax2.set_ylabel('Syringe fill level (mm)')
    ax2.set_xlabel('Timestamp')
    ax1.plot(pd.to_datetime(rfc_data['Timestamp']), rfc_data['Height (mm)'])
    ax2.plot(pd.to_datetime(sp_data['Timestamp']), sp_data['SP Position (mL)'])
    fig.autofmt_xdate()

    # plt.plot(pd.to_datetime(sp_data['Timestamp']), sp_data[' SP Position (mL)'])
    # plt.plot(pd.to_datetime(rfc_data['Timestamp']), rfc_data[" Frequency (Hz)"]/7500)
    plt.show()

    if savepath is not None:
        fig.savefig(os.path.join(r'../data_files', savepath))

        sp = os.path.join(r'../data_files', savepath.strip(".png")+".xlsx")

        wb = Workbook()
        sheet1 = wb.active
        sheet1.title = "SP data"
        for r in dataframe_to_rows(sp_data, index=True, header=True):
            sheet1.append(r)
        sheet2 = wb.create_sheet("RFC data")
        for r in dataframe_to_rows(rfc_data, index=True, header=True):
            sheet2.append(r)

        wb.save(filename=sp)

    return get_rough_maxmin(sp_data, rfc_data)


def collate_igordata(folder):
    xlsxfiles, igorfiles = get_all_xlsx_fnames(folder)
    print("{} files to process".format(len(xlsxfiles)))

    for x, i in zip(xlsxfiles, igorfiles):
        print(x, i)
        i_data = pd.read_csv(os.path.join(folder, i), header=None)

        wb = load_workbook(os.path.join(folder, x))
        ws = wb["RFC data"]
        ws["E1"] = "LVDT (V)"
        for r in range(6000):
            d = float(i_data.loc[r])
            ws["E"+str(r+3)] = d

        newname = x.strip(".xlsx") + "_LVDT.xlsx"
        wb.save(os.path.join(folder, newname))


def get_rough_maxmin(sp_df, rfc_df):
    spA = 0.5*(max(sp_df['SP Position (mL)'][800:1500]) - min(sp_df['SP Position (mL)'][800:1500]))
    rfcA = 0.5*(max(rfc_df['Height (mm)'][2000:4000]) - min(rfc_df['Height (mm)'][2000:4000]))
    lvdtA = 0.5*(max(rfc_df['LVDT (V)'][2000:4000]) - min(rfc_df['LVDT (V)'][2000:4000]))

    print(spA, rfcA, lvdtA)

    return spA, rfcA, lvdtA


def plot_roughmaxmin(flows, RFCamplitudes, LVDTamplitudes):
    plt.scatter(flows, RFCamplitudes, label="RFC")
    plt.scatter(flows, LVDTamplitudes, label="LVDT")

    plt.ylabel('Rough amplitude of motion (mm)')
    plt.xlabel('Flow at syringe pump (mL/min)')

    plt.legend()

    plt.tight_layout()
    plt.show()


def plot_log_maxmin(flows, RFCamplitudes, LVDTamplitudes, optflo):
    relflo = []
    for flo in flows:
        relflo.append(flo / optflo)

    plt.scatter(relflo, RFCamplitudes, label="RFC")
    plt.scatter(relflo, LVDTamplitudes, label="LVDT")

    plt.ylabel('Rough amplitude of motion (mm)')
    plt.xlabel('Flow at syringe pump (mL/min)')

    plt.legend()

    plt.tight_layout()
    plt.show()


def do_rough_maxmin(folder):
    fnames = get_all_LVDTxlsx_fnames(folder)

    flows = []
    RFCamplitudes, LVDTamplitudes = [], []

    wb = Workbook()
    ws = wb.active
    ws.append(["Flow (mL/min)", "RFC amplitude (mm)", "LVDT amplitude (V)"])

    for fname in fnames:
        code = fname.split("_")
        flows.append(float(code[2]))
        sp1, rfc1 = read_in_data(folder, fname)

        sp, rfcA, lvdtA = get_rough_maxmin(sp1, rfc1)

        RFCamplitudes.append(rfcA)
        LVDTamplitudes.append(lvdtA)

        ws.append([float(code[2]), rfcA, lvdtA])

    savepath = os.path.join(folder, "RoughMaxMin.xlsx")

    wb.save(savepath)
    print(savepath)

    plot_log_maxmin(flows, RFCamplitudes, LVDTamplitudes, 60*1.15)


def plot_timeseries_data(sp_df, rfc_df, savepath=None):
    fig, (ax1, ax2) = plt.subplots(2, sharex=True)
    # fig.suptitle('Piston height over time')
    ax1.set_ylabel('Height (mm)')
    # ax1.set_title('From Capacitor')
    ax2.set_ylabel('Syringe fill level (mm)')
    ax2.set_xlabel('Timestamp')
    ax1.plot(pd.to_datetime(rfc_df['Timestamp']), rfc_df['Height (mm)']-np.average(rfc_df['Height (mm)']), label="RFC")
    ax1.plot(pd.to_datetime(rfc_df['Timestamp']), rfc_df['LVDT (mm)']-np.average(rfc_df['LVDT (mm)']), label="LVDT")
    ax2.plot(pd.to_datetime(sp_df['Timestamp']), sp_df['SP Position (mL)'], label="SP")
    fig.autofmt_xdate()
    ax1.legend()

    # plt.show()
    if savepath is not None:
        plt.tight_layout()
        plt.savefig(savepath)

    plt.close(fig=fig)


def combine_lvdt_rfc_files(mydir):
    # get both lvdt and rfc data files from Igor Pro
    files = []
    igor_data = []

    for file in os.listdir(mydir):
        if file.startswith("LH_hgt") and file.endswith(".csv"):
            files.append(file)
        if "dP_volt" in file:
            igor_data.append(file)

    return files, igor_data


def combine_lvdt_rfc_data(folder):
    # e.g. use with
    # folder = r'I:\MSL\Shared\MSL Kibble Balance\_p_PressureManifoldConsiderations\Flow and Pressure control\SyringePumpTests\20201023_RFC-LVDT_IgorData'
    rfcf, lvdtf = combine_lvdt_rfc_files(folder)

    all_rfc_data = pd.DataFrame()
    all_lvdt_data = pd.DataFrame()

    for rf, lf in zip(rfcf[4:], lvdtf[4:]):
        rfc_data = pd.read_csv(os.path.join(folder, rf), header=None)
        rfc_data.columns = ["RF"]
        # print(rfc_data)
        all_rfc_data = all_rfc_data.append(rfc_data)

        lvdt_data = pd.read_csv(os.path.join(folder, lf), header=None)
        lvdt_data.columns = ["LVDT"]
        # print(lvdt_data)
        all_lvdt_data = all_lvdt_data.append(lvdt_data)

        # plt.scatter(rfc_data, lvdt_data)
        # plt.savefig(os.path.join(folder, rf.strip("csv")+"png"))

        # print(rf, lf, float(rfc_data.loc[50])/float(lvdt_data.loc[50]))

    plt.scatter(all_rfc_data, all_lvdt_data)
    plt.xlabel("Frequency (Hz)")
    plt.ylabel("LVDT Voltage (V)")
    plt.tight_layout()
    plt.savefig(os.path.join(folder, "CollatedData"))
    plt.show()

    all_data = pd.concat([all_rfc_data, all_lvdt_data], axis=1)
    all_data.to_excel(os.path.join(folder, "CollatedData.xlsx"))


def get_LVDTcal_from_RFCcal(datafile):
    # used with
    # folder = r'I:\MSL\Shared\MSL Kibble Balance\_p_PressureManifoldConsiderations\Flow and Pressure control\SyringePumpTests\20201023_RFC-LVDT_IgorData'
    # fname = "CollatedData.xlsx"

    df = pd.read_excel(datafile)
    # headers = list(df.columns.values.tolist())

    cal_height_rf = convert_cal_rf(df["RF"])  # headers[1]; get RFC data as Hz and convert it to position in mm

    plt.scatter(df["LVDT"], cal_height_rf)  # plot LVDT (headers[2]) against RFC data as position in mm
    plt.show()
    print(fit_quadratic(df["LVDT"], cal_height_rf, 0.1, 1, 0.1))

    # returns calibration [a, b, c] parameters: [-0.03542046  1.07974847 -0.39624465]
    # linear fit returns [ 0.82622535 -0.01726006]; Standard deviations: [0.00022915 0.00075441] but not a great fit.


def plot_rfc_lvdt():
    folder = r'I:\MSL\Shared\MSL Kibble Balance\_p_PressureManifoldConsiderations\Flow and Pressure control\SyringePumpTests\20201023_RFC-LVDT_IgorData'

    rf = "LH_hgt_scan0009.csv"
    lf = "dP_volt_scan0009.csv"

    rfc_data = pd.read_csv(os.path.join(folder, rf), header=None)
    rfc_data.columns = ["RF"]
    print(rfc_data)

    lvdt_data = pd.read_csv(os.path.join(folder, lf), header=None)
    lvdt_data.columns = ["LVDT"]
    print(lvdt_data)

    plt.plot(convert_cal_rf(rfc_data["RF"]), label="RF")
    plt.plot(convert_cal_lvdt(lvdt_data["LVDT"]), label="LVDT")
    plt.xlabel("Measurement number")
    # plt.xlabel("Frequency (Hz) converted to Height (mm)")
    plt.ylabel("Height (mm)")
    plt.tight_layout()
    plt.savefig(os.path.join(folder, lf.strip(".csv")+"_overTime.png"))

    plt.show()


if __name__ == "__main__":
    folder_0p5 = r"C:\Users\r.hawke\PycharmProjects\CetoniSP\data_files\2020-10-23 TriangleWaves 0.5mL"
    folder_1Hz = r"C:\Users\r.hawke\PycharmProjects\CetoniSP\data_files\2020-10-23 TriangleWaves 1Hz"
    fname = "Tri_0.1_12_1603423105.3048003_LVDT.xlsx"

    folder_20mL = r"G:\Shared drives\MSL - Shared\MSL Kibble Balance\_p_PressureManifoldConsiderations\Flow and Pressure control\SyringePumpTests\20201105 TriangleWaves 0.2mL"
    folder_0p05 = r"G:\Shared drives\MSL - Shared\MSL Kibble Balance\_p_PressureManifoldConsiderations\Flow and Pressure control\SyringePumpTests\20201109 TriangleWaves 0.05mL"

    fol_steps = r'G:\Shared drives\MSL - Shared\MSL Kibble Balance\_p_PressureManifoldConsiderations\Flow and Pressure control\SyringePumpTests\20201201_Step_loflo'

    fol = r"G:\Shared drives\MSL - Shared\MSL Kibble Balance\_p_PressureManifoldConsiderations\Flow and Pressure control\SyringePumpTests\20201209 Steps 0.75mL"
    collate_sp_rfc_lvdt(fol)


    # for fname in get_all_fnames(fol, pattern="LVDT_", endpattern='.csv'):
    #     i_data = pd.read_csv(os.path.join(fol, fname), header=None)
    #     i_data.columns = ['LVDT (V)']
    #     x = np.linspace(0, 999, num=1000)
    #     pars, stdres = fit_linear(x, i_data['LVDT (V)'], a=0, b=0.7)
    #     print(fname, *pars, stdres)


    # get_LVDTcal_from_RFCcal(os.path.join(fol, f1))



    # flows = [10, 20, 30, 40, 50, 45, 35, 27.5, 25, 22.5, 15, 5]
    # amps = []
    # for spf, rff in zip(sp_files, rf_files):
    #     sp_csv = os.path.join(folder_20mL, spf)
    #     rf_csv = os.path.join(folder_20mL, rff)
    #     a, b = plot_sp_rfc_data(sp_csv, rf_csv)
    #     amps.append(b)


    # plot_roughmaxmin(flows, amps, LVDTamplitudes=None)
    # sp1, df2 = read_in_data(folder, fname)

    # rfcA = convert_cal_rf(df2['Frequency (Hz)'][2000:4000])
    # rfcB = df2['Height (mm)'][2000:4000]
    # lvdtA = df2['LVDT (V)'][2000:4000]
    # lvdtB = convert_cal_lvdt(lvdtA, offset=-3.9)
    #
    # plt.plot(rfcB, label="RF_asHeight")
    # plt.plot(rfcA, label="RF_fromFreq")
    # plt.plot(lvdtA, label="LVDT_V")
    # plt.plot(lvdtB, label="LVDT_asHeight")
    # plt.legend()
    # plt.show()

    # do_rough_maxmin(folder_0p05)


